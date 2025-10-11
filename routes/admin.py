"""
Admin Routes Module

This module contains all admin-specific routes for the Panchayat Management System.
It handles dashboard, record management, user management, and department/scheme management.
"""

from flask import Blueprint, render_template, request, redirect, url_for, flash, session, make_response, jsonify
from models.admin import PanchayatRecord
from models.login import User
from models.department import Department
from models.scheme import Scheme
from routes.login import admin_required, login_required, superadmin_required, can_edit_records
from datetime import datetime, timedelta
from bson import ObjectId
import pandas as pd
import io
import json

# Initialize admin blueprint
admin_bp = Blueprint('admin', __name__)
mongo = None

def intialize_admin(db):
    """Initialize MongoDB connection for admin routes"""
    global mongo
    mongo = db

@admin_bp.route("/dashboard")
@login_required
@admin_required
def dashboard():
    try:
        # Use PanchayatRecord model for statistics
        stats = PanchayatRecord.get_statistics(mongo)
        users_count = mongo.db.users.count_documents({})
        
        # Additional chart data
        # House status distribution
        status_pipeline = [
            {"$group": {"_id": "$house_status", "count": {"$sum": 1}}},
            {"$sort": {"count": -1}}
        ]
        status_stats = list(mongo.db.panchayat_records.aggregate(status_pipeline))
        
        # Monthly trends (last 6 months)
        from datetime import datetime, timedelta
        import calendar
        
        six_months_ago = datetime.now() - timedelta(days=180)
        monthly_pipeline = [
            {"$match": {"created_at": {"$gte": six_months_ago}}},
            {"$group": {
                "_id": {
                    "year": {"$year": "$created_at"},
                    "month": {"$month": "$created_at"}
                },
                "count": {"$sum": 1}
            }},
            {"$sort": {"_id.year": 1, "_id.month": 1}}
        ]
        monthly_stats = list(mongo.db.panchayat_records.aggregate(monthly_pipeline))
        
        # Format monthly data for chart
        monthly_labels = []
        monthly_data = []
        for stat in monthly_stats:
            month_name = calendar.month_abbr[stat['_id']['month']]
            monthly_labels.append(month_name)
            monthly_data.append(stat['count'])

        # Calculate tab-specific statistics with error handling
        try:
            approved_stats = calculate_approved_stats(mongo)
        except Exception as e:
            approved_stats = {'count': 0, 'recent_count': 0, 'total_amount': 0, 'avg_amount': 0, 'this_month': 0, 'completion_rate': 0, 'avg_processing_days': 0, 'fastest_approval': 0}
        
        try:
            pending_stats = calculate_pending_stats(mongo)
        except Exception as e:
            pending_stats = {'count': 0, 'recent_count': 0, 'overdue_count': 0, 'avg_waiting_days': 0, 'longest_waiting': 0, 'total_amount': 0, 'avg_amount': 0}
        
        try:
            inreview_stats = calculate_inreview_stats(mongo)
        except Exception as e:
            inreview_stats = {'count': 0, 'today_count': 0, 'reviewers': 0, 'avg_per_reviewer': 0, 'avg_review_days': 0, 'efficiency': 0, 'doc_verified': 0, 'field_pending': 0, 'final_pending': 0}
        
        try:
            disapproved_stats = calculate_disapproved_stats(mongo)
        except Exception as e:
            disapproved_stats = {'count': 0, 'this_month': 0, 'resubmitted': 0, 'resubmit_rate': 0, 'top_reason_count': 0, 'top_reason': 'Main Issue', 'disapproval_rate': 0, 'trend': 'stable', 'trend_text': 'vs last month'}
        
        try:
            rejected_stats = calculate_rejected_stats(mongo)
        except Exception as e:
            rejected_stats = {'count': 0, 'recent_count': 0}
        
        return render_template("admin/dashboard.html", 
                             users_count=users_count,
                             records_count=stats['total_records'],
                             total_amount=stats['total_amount'],
                             category_stats=stats['category_stats'],
                             panchayat_stats=stats['panchayat_stats'],
                             status_stats=status_stats,
                             monthly_labels=monthly_labels,
                             monthly_data=monthly_data,
                             # Add tab-specific stats
                             approved_status=approved_stats,  # Fix: Use approved_status not approved_stats
                             approved_stats=approved_stats,
                             pending_stats=pending_stats,
                             inreview_stats=inreview_stats,
                             disapproved_stats=disapproved_stats,
                             rejected_stats=rejected_stats)
                             
    except Exception as e:
        flash('An error occurred while loading the dashboard.', 'error')
        # Return with default empty stats
        empty_stats = get_default_stats()
        return render_template("admin/dashboard.html", 
                             users_count=0, 
                             records_count=0, 
                             total_amount=0,
                             category_stats=[], 
                             panchayat_stats=[],
                             status_stats=[], 
                             monthly_labels=[], 
                             monthly_data=[],
                             approved_status=empty_stats,
                             approved_stats=empty_stats,
                             pending_stats=empty_stats,
                             inreview_stats=empty_stats,
                             disapproved_stats=empty_stats,
                             rejected_stats=empty_stats)

def calculate_approved_stats(mongo):
    """Calculate statistics for approved records (Complete status)"""
    try:
        # Total approved count (Complete houses)
        total_count = mongo.db.panchayat_records.count_documents({"house_status": "Complete"})
        
        # This month count
        start_of_month = datetime.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        this_month = mongo.db.panchayat_records.count_documents({
            "house_status": "Complete",
            "created_at": {"$gte": start_of_month}
        })
        
        # Total amount for approved records
        amount_pipeline = [
            {"$match": {"house_status": "Complete"}},
            {"$group": {"_id": None, "total": {"$sum": {"$toDouble": "$amount_released"}}}}
        ]
        amount_result = list(mongo.db.panchayat_records.aggregate(amount_pipeline))
        total_amount = amount_result[0]['total'] if amount_result else 0
        
        # Calculate average processing days (dummy calculation)
        avg_processing_days = 15
        
        return {
            'count': total_count,
            'this_month': this_month,
            'recent_count': this_month,
            'total_amount': total_amount,
            'avg_amount': round((total_amount / total_count) if total_count > 0 else 0, 2),
            'avg_processing_days': avg_processing_days,
            'fastest_approval': 7,
            'completion_rate': 85
        }
    except Exception as e:
        print(f"Error calculating approved stats: {e}")
        return get_default_stats()

def calculate_pending_stats(mongo):
    """Calculate statistics for pending records (Under Construction)"""
    try:
        # Count pending records
        total_count = mongo.db.panchayat_records.count_documents({"house_status": "Under Construction"})
        
        # Today's count
        today_start = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        today_count = mongo.db.panchayat_records.count_documents({
            "house_status": "Under Construction",
            "created_at": {"$gte": today_start}
        })
        
        # Overdue count (records older than 30 days)
        thirty_days_ago = datetime.now() - timedelta(days=30)
        overdue_count = mongo.db.panchayat_records.count_documents({
            "house_status": "Under Construction",
            "created_at": {"$lt": thirty_days_ago}
        })
        
        # Total amount for pending records
        amount_pipeline = [
            {"$match": {"house_status": "Under Construction"}},
            {"$group": {"_id": None, "total": {"$sum": {"$toDouble": "$amount_released"}}}}
        ]
        amount_result = list(mongo.db.panchayat_records.aggregate(amount_pipeline))
        total_amount = amount_result[0]['total'] if amount_result else 0
        
        return {
            'count': total_count,
            'recent_count': today_count,
            'overdue_count': overdue_count,
            'total_amount': total_amount,
            'avg_waiting_days': 22,
            'longest_waiting': 45,
            'avg_amount': 15
        }
    except Exception as e:
        print(f"Error calculating pending stats: {e}")
        return get_default_stats()

def calculate_inreview_stats(mongo):
    """Calculate statistics for records in review (Incomplete status)"""
    try:
        # Count in-review records
        total_count = mongo.db.panchayat_records.count_documents({"house_status": "Incomplete"})
        
        # Today's count
        today_start = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        today_count = mongo.db.panchayat_records.count_documents({
            "house_status": "Incomplete",
            "created_at": {"$gte": today_start}
        })
        
        return {
            'count': total_count,
            'today_count': today_count,
            'reviewers': 5,
            'avg_per_reviewer': round(total_count / 5) if total_count > 0 else 0,
            'avg_review_days': 8,
            'efficiency': 15,
            'doc_verified': round(total_count * 0.6),
            'field_pending': round(total_count * 0.3),
            'final_pending': round(total_count * 0.1)
        }
    except Exception as e:
        print(f"Error calculating inreview stats: {e}")
        return get_default_stats()

def calculate_disapproved_stats(mongo):
    """Calculate statistics for disapproved records"""
    try:
        # Using priority >= 8 as disapproved for demo
        total_count = mongo.db.panchayat_records.count_documents({"priority": {"$gte": 8}})
        
        # This month count
        start_of_month = datetime.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        this_month = mongo.db.panchayat_records.count_documents({
            "priority": {"$gte": 8},
            "created_at": {"$gte": start_of_month}
        })
        
        return {
            'count': total_count,
            'this_month': this_month,
            'resubmitted': round(total_count * 0.3),
            'resubmit_rate': 30,
            'top_reason': 'Incomplete Documentation',
            'top_reason_count': round(total_count * 0.4),
            'disapproval_rate': 12,
            'trend': 'down',
            'trend_text': '↓ 5% vs last month'
        }
    except Exception as e:
        print(f"Error calculating disapproved stats: {e}")
        return get_default_stats()

def calculate_rejected_stats(mongo):
    """Calculate statistics for rejected records"""
    try:
        # Using priority == 10 as rejected for demo
        total_count = mongo.db.panchayat_records.count_documents({"priority": 10})
        
        return {
            'count': total_count,
            'this_month': round(total_count * 0.1),
            'total_amount': total_count * 50000,
            'rejection_rate': 5,
            'top_reason': 'Eligibility Criteria Not Met',
            'top_reason_count': round(total_count * 0.6)
        }
    except Exception as e:
        print(f"Error calculating rejected stats: {e}")
        return get_default_stats()

def get_default_stats():
    """Return default stats when there's an error or no data"""
    return {
        'count': 0,
        'this_month': 0,
        'recent_count': 0,
        'total_amount': 0,
        'avg_amount': 0,
        'avg_processing_days': 0,
        'fastest_approval': 0,
        'completion_rate': 0,
        'overdue_count': 0,
        'avg_waiting_days': 0,
        'longest_waiting': 0,
        'today_count': 0,
        'reviewers': 0,
        'avg_per_reviewer': 0,
        'avg_review_days': 0,
        'efficiency': 0,
        'doc_verified': 0,
        'field_pending': 0,
        'final_pending': 0,
        'resubmitted': 0,
        'resubmit_rate': 0,
        'top_reason': 'N/A',
        'top_reason_count': 0,
        'disapproval_rate': 0,
        'trend': 'stable',
        'trend_text': 'No change',
        'rejection_rate': 0
    }

@admin_bp.route('add-record', methods=['GET', 'POST'])
@login_required
@admin_required
def add_record():
    if request.method == 'POST':
        try:
            # Get scheme ID and validate access
            scheme_id = request.form.get('scheme_id')
            if not scheme_id:
                flash('Please select a scheme', 'error')
                return redirect(url_for('admin.add_record'))
            
            # Check if user has access to this scheme
            if not User.can_access_scheme(mongo, session.get('user_id'), scheme_id):
                flash('You do not have access to this scheme', 'error')
                return redirect(url_for('admin.add_record'))
            
            # Get scheme details for validation
            scheme_result = Scheme.get_scheme_by_id(mongo, scheme_id)
            if not scheme_result['success']:
                flash('Invalid scheme selected', 'error')
                return redirect(url_for('admin.add_record'))
            
            scheme = scheme_result['scheme']
            department_id = scheme['department_id']
            
            # Get scheme attributes to build custom data
            scheme_attributes = scheme.get('attributes', [])
            custom_data = {}
            
            # Process each attribute from the scheme
            for attr in scheme_attributes:
                field_name = attr.get('name')
                data_type = attr.get('type', 'string')
                field_value = request.form.get(field_name)
                
                if field_value:
                    # Convert data types based on scheme definition
                    if data_type == 'int':
                        try:
                            custom_data[field_name] = int(field_value)
                        except ValueError:
                            custom_data[field_name] = 0
                    elif data_type == 'float':
                        try:
                            custom_data[field_name] = float(field_value)
                        except ValueError:
                            custom_data[field_name] = 0.0
                    elif data_type == 'date':
                        try:
                            custom_data[field_name] = datetime.strptime(field_value, '%Y-%m-%d')
                        except ValueError:
                            custom_data[field_name] = None
                    elif data_type == 'boolean':
                        custom_data[field_name] = field_value == 'true'
                    else:  # string, enum, etc.
                        custom_data[field_name] = field_value
                else:
                    # Set default values based on data type
                    if data_type == 'int':
                        custom_data[field_name] = 0
                    elif data_type == 'float':
                        custom_data[field_name] = 0.0
                    elif data_type == 'date':
                        custom_data[field_name] = None
                    elif data_type == 'boolean':
                        custom_data[field_name] = False
                    else:
                        custom_data[field_name] = ''
            
            # Prepare record data with scheme and department info
            record_data = {
                'department_id': department_id,
                'scheme_id': scheme_id,
                'custom_data': custom_data
            }

            # Use PanchayatRecord model to create record
            result = PanchayatRecord.create_record(
                mongo,
                record_data,
                created_by= session.get('username'))
            
            if result['success']:
                flash('Record added successfully!', 'success')
                return redirect(url_for('admin.view_records'))
            else:
                flash(result['message'], 'error')
            
        except Exception as e:
            flash('An error occurred while adding the record.', 'error')

    # GET request - show form
    return render_template('admin/add_record.html')

@admin_bp.route('/view-records')
@login_required
@admin_required
def view_records():
    try:
        page = int(request.args.get('page', 1))
        search = request.args.get('search', '')
        per_page = 10
        
        # Get filter parameters
        department_ids = request.args.getlist('department_ids')
        scheme_ids = request.args.getlist('scheme_ids')
        
        # Get user's access information
        user_id = session.get('user_id')
        user_access = User.get_user_access_info(mongo, user_id)

        
        # Use PanchayatRecord model to get records with user access filtering
        result = PanchayatRecord.get_all_records(
            mongo,
            page=page,
            per_page=per_page,
            search=search, 
            department_ids=department_ids if department_ids else None, 
            scheme_ids=scheme_ids if scheme_ids else None,
            user_id=user_id
        )
        
        if result['success']:
            # Get dashboard statistics for the template
            try:
                stats = PanchayatRecord.get_statistics(mongo)
                approved_stats = calculate_approved_stats(mongo)
                pending_stats = calculate_pending_stats(mongo)
                inreview_stats = calculate_inreview_stats(mongo)
                disapproved_stats = calculate_disapproved_stats(mongo)
                rejected_stats = calculate_rejected_stats(mongo)
            except Exception as e:
                # Provide default values
                stats = {'total_records': 0, 'total_amount': 0, 'category_stats': [], 'panchayat_stats': []}
                approved_stats = {'count': 0, 'recent_count': 0}
                pending_stats = {'count': 0, 'recent_count': 0}
                inreview_stats = {'count': 0, 'recent_count': 0}
                disapproved_stats = {'count': 0, 'recent_count': 0}
                rejected_stats = {'count': 0, 'recent_count': 0}
            
            return render_template('admin/view_records.html', 
                                 records=result['records'],
                                 page=result['page'],
                                 total_records=result['total_records'],
                                 total_pages=result['total_pages'],
                                 search=search,
                                 department_ids=department_ids,
                                 scheme_ids=scheme_ids,
                                 user_access=user_access,
                                 can_edit=session.get('is_superadmin', False),
                                 approved_status=approved_stats,
                                 approved_stats=approved_stats,
                                 pending_stats=pending_stats,
                                 inreview_stats=inreview_stats,
                                 disapproved_stats=disapproved_stats,
                                 rejected_stats=rejected_stats,
                                 records_count=stats['total_records'],
                                 total_amount=stats['total_amount'])
        else:
            flash('Error loading records', 'error')
            return render_template('admin/view_records.html', records=[], page=1, 
                                 total_records=0, total_pages=0, search='', 
                                 department_ids=[], scheme_ids=[],
                                 user_access=user_access,
                                 can_edit=session.get('is_superadmin', False),
                                 approved_status={'count': 0, 'recent_count': 0},
                                 approved_stats={'count': 0, 'recent_count': 0},
                                 pending_stats={'count': 0, 'recent_count': 0},
                                 inreview_stats={'count': 0, 'recent_count': 0},
                                 disapproved_stats={'count': 0, 'recent_count': 0},
                                 rejected_stats={'count': 0, 'recent_count': 0},
                                 records_count=0, total_amount=0)
            
    except Exception as e:
        flash('Error loading records', 'error')
        return render_template('admin/view_records.html', records=[], page=1, 
                             total_records=0, total_pages=0, search='', 
                             department_ids=[], scheme_ids=[],
                             user_access=None,
                             can_edit=session.get('is_superadmin', False),
                             approved_status={'count': 0, 'recent_count': 0},
                             approved_stats={'count': 0, 'recent_count': 0},
                             pending_stats={'count': 0, 'recent_count': 0},
                             inreview_stats={'count': 0, 'recent_count': 0},
                             disapproved_stats={'count': 0, 'recent_count': 0},
                             rejected_stats={'count': 0, 'recent_count': 0},
                             records_count=0, total_amount=0)

@admin_bp.route('/export-excel')
@login_required
@admin_required
def export_excel():
    try:
        # Get filter parameters
        department_ids = request.args.getlist('department_ids')
        scheme_ids = request.args.getlist('scheme_ids')
        
        # Use PanchayatRecord model for filtered export data
        export_data = PanchayatRecord.export_to_excel_data_filtered(
            mongo, session.get('user_id'), 
            department_ids if department_ids else None,
            scheme_ids if scheme_ids else None
        )
        
        if not export_data:
            flash('No data available for export', 'warning')
            return redirect(url_for('admin.dashboard'))

        df = pd.DataFrame(export_data)

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Panchayat Data')
        
        output.seek(0)

        response = make_response(output.getvalue())
        response.headers['Content-Disposition'] = f'attachment; filename=panchayat_data_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        
        return response
    except Exception as e:
        flash('An error occurred while exporting data.', 'error')
        return redirect(url_for('admin.dashboard'))

@admin_bp.route('/delete-record/<record_id>', methods=['POST'])
@login_required
@admin_required
def delete_record(record_id):
    try:
        # Use PanchayatRecord model to delete record
        success = PanchayatRecord.delete_record(
            mongo,
            record_id,
            deleted_by=session.get('username'))
        
        if success:
            return jsonify({'success': True, 'message': 'Record deleted successfully!'})
        else:
            return jsonify({'success': False, 'message': 'Record not found or could not be deleted'})
            
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error deleting record: {str(e)}'})

@admin_bp.route('/api/update-record/<record_id>', methods=['POST'])
@login_required
@admin_required
def update_record_api(record_id):
    try:
        # Get the record to understand its structure
        record = PanchayatRecord.get_record_by_id(mongo, record_id)
        if not record:
            return jsonify({'success': False, 'message': 'Record not found'})
        
        # Get scheme information to understand the data structure
        scheme_id = record.get('scheme_id')
        # Handle ObjectId format for scheme_id
        if scheme_id:
            if isinstance(scheme_id, dict) and '$oid' in scheme_id:
                scheme_id = scheme_id['$oid']
            scheme_result = Scheme.get_scheme_by_id(mongo, scheme_id)
            scheme = scheme_result.get('scheme') if scheme_result.get('success') else None
        else:
            scheme = None
        
        # Get update data from request
        update_data = request.get_json()
        
        # Prepare update data with custom_data structure
        final_update_data = {
            'updated_by': session.get('username'),
            'updated_at': datetime.utcnow()
        }
        
        # Process custom_data fields based on scheme attributes
        if scheme and scheme.get('attributes'):
            custom_data = {}
            for attr in scheme['attributes']:
                field_name = attr['name']
                field_type = attr['type']
                field_value = update_data.get(field_name)
                
                if field_value is not None and field_value != '':
                    # Convert value based on field type
                    if field_type == 'int':
                        custom_data[field_name] = int(field_value)
                    elif field_type == 'float':
                        custom_data[field_name] = float(field_value)
                    elif field_type == 'date':
                        custom_data[field_name] = datetime.strptime(field_value, '%Y-%m-%d')
                    elif field_type == 'boolean':
                        custom_data[field_name] = field_value.lower() in ['true', '1', 'yes', 'on']
                    else:  # string, enum
                        custom_data[field_name] = field_value
                else:
                    custom_data[field_name] = None
            
            final_update_data['custom_data'] = custom_data
        
        # Update record in database
        success = PanchayatRecord.update_record(
            mongo,
            record_id,
            final_update_data,
            updated_by=session.get('username')
        )
        
        if success:
            return jsonify({'success': True, 'message': 'Record updated successfully!'})
        else:
            return jsonify({'success': False, 'message': 'Failed to update record'})
            
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error updating record: {str(e)}'})

@admin_bp.route('/manage-users')
@login_required
@superadmin_required
def manage_users():
    try:
        page = int(request.args.get('page', 1))
        search = request.args.get('search', '')
        per_page = 10
        
        # Get all users with pagination
        query = {}
        if search:
            query = {
                '$or': [
                    {'username': {'$regex': search, '$options': 'i'}},
                    {'email': {'$regex': search, '$options': 'i'}},
                    {'role': {'$regex': search, '$options': 'i'}}
                ]
            }
        
        total_users = mongo.db.users.count_documents(query)
        total_pages = (total_users + per_page - 1) // per_page
        
        users = list(mongo.db.users.find(query)
                    .sort('created_at', -1)
                    .skip((page - 1) * per_page)
                    .limit(per_page))
        
        # Get departments and schemes for user creation form
        dept_result = Department.get_all_departments(mongo, active_only=True)
        departments = dept_result['departments'] if dept_result['success'] else []
        
        scheme_result = Scheme.get_all_schemes(mongo, active_only=True)
        schemes = scheme_result['schemes'] if scheme_result['success'] else []
        
        return render_template('admin/manage_users.html', 
                             users=users,
                             page=page,
                             total_users=total_users,
                             total_pages=total_pages,
                             search=search,
                             departments=departments,
                             schemes=schemes)
        
    except Exception as e:
        flash('Error loading users', 'error')
        return render_template('admin/manage_users.html', users=[], page=1, 
                             total_users=0, total_pages=0, search='', departments=[], schemes=[])

@admin_bp.route('/edit-user/<user_id>', methods=['GET', 'POST'])
@login_required
@superadmin_required
def edit_user(user_id):
    if request.method == 'POST':
        try:
            new_role = request.form.get('role')
            is_admin = new_role in ['admin', 'superadmin']
            is_superadmin = new_role == 'superadmin'
            
            # Prevent superadmin from demoting themselves
            if user_id == session.get('user_id') and not is_superadmin:
                flash('You cannot demote yourself from superadmin role.', 'error')
                return redirect(url_for('admin.manage_users'))
            
            # Get department and scheme access from checkboxes
            department_access = request.form.getlist('department_access')
            scheme_access = request.form.getlist('scheme_access')

            # Handle 'all' selections
            if 'all' in department_access:
                department_access = ['all']
            elif not department_access:
                department_access = []

            if 'all' in scheme_access:
                scheme_access = ['all']
            elif not scheme_access:
                scheme_access = []
            
            
            update_data = {
                'role': new_role,
                'email': request.form.get('email', '').strip() or None,
                'is_admin': is_admin,
                'is_superadmin': is_superadmin,
                'is_active': request.form.get('is_active') == 'on',
                'department_access': department_access,
                'scheme_access': scheme_access,
                'updated_at': datetime.utcnow()
            }
            
            success = User.update_user(
                mongo,
                user_id,
                update_data,
                updated_by=session.get('username')
            )
            
            if success:
                flash('User updated successfully!', 'success')
                return redirect(url_for('admin.manage_users'))
            else:
                flash('Failed to update user', 'error')
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            flash('An error occurred while updating user.', 'error')
    
    # GET request - show edit form
    try:
        user = User.get_user_by_id(mongo, user_id)
        if not user:
            flash('User not found', 'error')
            return redirect(url_for('admin.manage_users'))
        
        # Ensure department_access and scheme_access are lists
        if 'department_access' not in user or user['department_access'] is None:
            user['department_access'] = []
        elif not isinstance(user['department_access'], list):
            user['department_access'] = [user['department_access']]
            
        if 'scheme_access' not in user or user['scheme_access'] is None:
            user['scheme_access'] = []
        elif not isinstance(user['scheme_access'], list):
            user['scheme_access'] = [user['scheme_access']]
        
        # Get all departments and schemes for selection
        dept_result = Department.get_all_departments(mongo, active_only=True)
        departments = dept_result['departments'] if dept_result['success'] else []
        
        scheme_result = Scheme.get_all_schemes(mongo, active_only=True)
        schemes = scheme_result['schemes'] if scheme_result['success'] else []
        
        return render_template('admin/edit_user.html', user=user, departments=departments, schemes=schemes)
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        flash('Error loading user', 'error')
        return redirect(url_for('admin.manage_users'))

@admin_bp.route('/toggle-user-status/<user_id>')
@login_required
@superadmin_required
def toggle_user_status(user_id):
    try:
        # Prevent superadmin from deactivating themselves
        if user_id == session.get('user_id'):
            flash('You cannot deactivate your own account.', 'error')
            return redirect(url_for('admin.manage_users'))
        
        user = User.get_user_by_id(mongo, user_id)
        if not user:
            flash('User not found', 'error')
            return redirect(url_for('admin.manage_users'))
        
        new_status = not user.get('is_active', True)
        update_data = {
            'is_active': new_status,
            'updated_at': datetime.utcnow()
        }
        
        success = User.update_user(
            mongo,
            user_id,
            update_data,
            updated_by=session.get('username')
        )
        
        if success:
            status_text = 'activated' if new_status else 'deactivated'
            flash(f'User {status_text} successfully!', 'success')
        else:
            flash('Failed to update user status', 'error')
        
    except Exception as e:
        flash('Error updating user status', 'error')
    
    return redirect(url_for('admin.manage_users'))

@admin_bp.route('/delete-user/<user_id>')
@login_required
@superadmin_required
def delete_user(user_id):
    try:
        # Prevent superadmin from deleting themselves
        if user_id == session.get('user_id'):
            flash('You cannot delete your own account.', 'error')
            return redirect(url_for('admin.manage_users'))
        
        from bson import ObjectId
        result = mongo.db.users.delete_one({'_id': ObjectId(user_id)})
        
        if result.deleted_count > 0:
            flash('User deleted successfully!', 'success')
        else:
            flash('User not found or could not be deleted', 'error')
        
    except Exception as e:
        flash('Error deleting user', 'error')
    
    return redirect(url_for('admin.manage_users'))

@admin_bp.route('/edit-record/<record_id>', methods=['GET', 'POST'])
@login_required
@can_edit_records
def edit_record(record_id):
    if request.method == 'POST':
        try:
            # Get the record to understand its structure
            record = PanchayatRecord.get_record_by_id(mongo, record_id)
            if not record:
                flash('Record not found', 'error')
                return redirect(url_for('admin.view_records'))
            
            # Get scheme information to understand the data structure
            scheme_id = record.get('scheme_id')
            scheme = Scheme.get_scheme_by_id(mongo, scheme_id) if scheme_id else None
            
            # Prepare update data with custom_data structure
            update_data = {
                'updated_by': session.get('username'),
                'updated_at': datetime.utcnow()
            }
            
            # Process custom_data fields based on scheme attributes
            if scheme and scheme.get('attributes'):
                custom_data = {}
                for attr in scheme['attributes']:
                    field_name = attr['name']
                    field_type = attr['type']
                    field_value = request.form.get(field_name)
                    
                    if field_value is not None and field_value != '':
                        # Convert value based on field type
                        if field_type == 'int':
                            custom_data[field_name] = int(field_value)
                        elif field_type == 'float':
                            custom_data[field_name] = float(field_value)
                        elif field_type == 'date':
                            custom_data[field_name] = datetime.strptime(field_value, '%Y-%m-%d')
                        elif field_type == 'boolean':
                            custom_data[field_name] = field_value.lower() in ['true', '1', 'yes', 'on']
                        else:  # string, enum
                            custom_data[field_name] = field_value
                    else:
                        custom_data[field_name] = None
                
                update_data['custom_data'] = custom_data
            
            # Update record
            success = PanchayatRecord.update_record(mongo, record_id, update_data)
            
            if success:
                flash('Record updated successfully!', 'success')
                return redirect(url_for('admin.view_records'))
            else:
                flash('Failed to update record', 'error')
            
        except Exception as e:
            flash('An error occurred while updating the record.', 'error')
    
    # GET request - show edit form
    try:
        record = PanchayatRecord.get_record_by_id(mongo, record_id)
        if not record:
            flash('Record not found', 'error')
            return redirect(url_for('admin.view_records'))
        
        # Get scheme information for form generation
        scheme_id = record.get('scheme_id')
        scheme = Scheme.get_scheme_by_id(mongo, scheme_id) if scheme_id else None
        
        return render_template('admin/edit_record.html', record=record, scheme=scheme)
        
    except Exception as e:
        flash('Error loading record', 'error')
        return redirect(url_for('admin.view_records'))

# Add these routes to your admin.py file

@admin_bp.route("/create_user", methods=["POST"])
@login_required
@superadmin_required
def create_user():
    try:
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        confirm_password = request.form.get('confirm_password', '')
        email = request.form.get('email', '').strip()
        role = request.form.get('role', '')
        status = request.form.get('status', 'active')
        full_name = request.form.get('full_name', '').strip()


        # Validation
        if not username or not password or not role:
            flash('Username, password, and role are required.', 'error')
            return redirect(url_for('admin.manage_users'))

        if len(username) < 3:
            flash('Username must be at least 3 characters long.', 'error')
            return redirect(url_for('admin.manage_users'))

        if len(password) < 6:
            flash('Password must be at least 6 characters long.', 'error')
            return redirect(url_for('admin.manage_users'))

        if password != confirm_password:
            flash('Passwords do not match.', 'error')
            return redirect(url_for('admin.manage_users'))

        if role not in ['user', 'admin', 'superadmin']:
            flash('Invalid role selected.', 'error')
            return redirect(url_for('admin.manage_users'))

        # Set role-based flags
        is_admin = role in ['admin', 'superadmin']
        is_superadmin = role == 'superadmin'

        # Get department and scheme access
        department_access_str = request.form.get('department_access', '').strip()
        scheme_access_str = request.form.get('scheme_access', '').strip()

        if department_access_str == 'all':
            department_access = ['all']
        elif department_access_str:
            department_access = [d.strip() for d in department_access_str.split(',') if d.strip()]
        else:
            department_access = []
    

        if scheme_access_str == 'all':
            scheme_access = ['all']
        elif scheme_access_str:
            scheme_access = [s.strip() for s in scheme_access_str.split(',') if s.strip()]
        else:
            scheme_access = []
        
        # Create user using the User model (dictionary approach)
        user_data = {
            'username': username,
            'password': password,
            'role': role,
            'email': email if email else None,
            'full_name': full_name if full_name else None,
            'is_active': status == 'active',
            'is_admin': role in ['admin', 'superadmin'],
            'is_superadmin': role == 'superadmin',
            'department_access': department_access,
            'scheme_access': scheme_access
        }

        result = User.create_user(mongo, user_data, created_by=session.get('username'))

        if result['success']:
            flash(f'User "{username}" created successfully with role "{role}".', 'success')
        else:
            flash(result['message'], 'error')

    except Exception as e:
        import traceback
        traceback.print_exc()
        flash('An error occurred while creating the user.', 'error')

    return redirect(url_for('admin.manage_users'))

# Department Management Routes

@admin_bp.route('/manage-departments')
@login_required
@superadmin_required
def manage_departments():
    """Manage departments - superadmin only"""
    try:
        result = Department.get_all_departments(mongo, active_only=False)
        if result['success']:
            departments = result['departments']
        else:
            departments = []
            flash('Error loading departments', 'error')
        
        # Get dashboard statistics for the template
        try:
            stats = PanchayatRecord.get_statistics(mongo)
            approved_stats = calculate_approved_stats(mongo)
            pending_stats = calculate_pending_stats(mongo)
            inreview_stats = calculate_inreview_stats(mongo)
            disapproved_stats = calculate_disapproved_stats(mongo)
            rejected_stats = calculate_rejected_stats(mongo)
        except Exception as e:
            print(f"Error loading dashboard stats: {e}")
            # Provide default values
            stats = {'total_records': 0, 'total_amount': 0, 'category_stats': [], 'panchayat_stats': []}
            approved_stats = {'count': 0, 'recent_count': 0}
            pending_stats = {'count': 0, 'recent_count': 0}
            inreview_stats = {'count': 0, 'recent_count': 0}
            disapproved_stats = {'count': 0, 'recent_count': 0}
            rejected_stats = {'count': 0, 'recent_count': 0}
        
        # Get the department ID to open from URL parameter
        open_department_id = request.args.get('open_department')
        
        return render_template('admin/manage_departments.html', 
                             departments=departments,
                             open_department_id=open_department_id,
                             approved_status=approved_stats,
                             approved_stats=approved_stats,
                             pending_stats=pending_stats,
                             inreview_stats=inreview_stats,
                             disapproved_stats=disapproved_stats,
                             rejected_stats=rejected_stats,
                             records_count=stats['total_records'],
                             total_amount=stats['total_amount'])
        
    except Exception as e:
        flash('Error loading departments', 'error')
        # Get the department ID to open from URL parameter
        open_department_id = request.args.get('open_department')
        
        return render_template('admin/manage_departments.html', 
                             departments=[],
                             open_department_id=open_department_id,
                             approved_status={'count': 0, 'recent_count': 0},
                             approved_stats={'count': 0, 'recent_count': 0},
                             pending_stats={'count': 0, 'recent_count': 0},
                             inreview_stats={'count': 0, 'recent_count': 0},
                             disapproved_stats={'count': 0, 'recent_count': 0},
                             rejected_stats={'count': 0, 'recent_count': 0},
                             records_count=0,
                             total_amount=0)

@admin_bp.route('/create-department', methods=['POST'])
@login_required
@superadmin_required
def create_department():
    """Create a new department"""
    try:
        department_data = {
            'name': request.form.get('name', '').strip(),
            'description': request.form.get('description', '').strip()
        }
        
        if not department_data['name']:
            flash('Department name is required', 'error')
            return redirect(url_for('admin.manage_departments'))
        
        result = Department.create_department(mongo, department_data,username=session.get('username'))
        
        if result['success']:
            flash('Department created successfully!', 'success')
            # Redirect with the new department ID to open its accordion
            return redirect(url_for('admin.manage_departments', open_department=result['department_id']))
        else:
            flash(result['message'], 'error')
            
    except Exception as e:
        flash('An error occurred while creating the department.', 'error')
    
    return redirect(url_for('admin.manage_departments'))

@admin_bp.route('/edit-department/<department_id>', methods=['GET', 'POST'])
@login_required
@superadmin_required
def edit_department(department_id):
    """Edit department"""
    if request.method == 'POST':
        try:
            update_data = {
                'name': request.form.get('name', '').strip(),
                'description': request.form.get('description', '').strip()
            }
            
            if not update_data['name']:
                flash('Department name is required', 'error')
                return redirect(url_for('admin.manage_departments'))
            
            result = Department.update_department(mongo, department_id, update_data,username=session.get('username'))
            
            if result['success']:
                flash('Department updated successfully!', 'success')
            else:
                flash(result['message'], 'error')
                
        except Exception as e:
            flash('An error occurred while updating the department.', 'error')
        
        return redirect(url_for('admin.manage_departments'))
    
    # GET request - show edit form
    try:
        department = Department.get_department_by_id(mongo, department_id)
        if not department:
            flash('Department not found', 'error')
            return redirect(url_for('admin.manage_departments'))
        
        return render_template('admin/edit_department.html', department=department)
        
    except Exception as e:
        flash('Error loading department', 'error')
        return redirect(url_for('admin.manage_departments'))

@admin_bp.route('/delete-department/<department_id>')
@login_required
@superadmin_required
def delete_department(department_id):
    """Delete department"""
    try:
        result = Department.delete_department(mongo, department_id,username=session.get('username'))
        
        if result['success']:
            flash('Department deleted successfully!', 'success')
        else:
            flash(result['message'], 'error')
            
    except Exception as e:
        flash('Error deleting department', 'error')
    
    return redirect(url_for('admin.manage_departments'))

# Scheme Management Routes

@admin_bp.route('/manage-schemes')
@login_required
@superadmin_required
def manage_schemes():
    """Manage schemes - superadmin only"""
    try:
        result = Scheme.get_all_schemes(mongo, active_only=False)
        if result['success']:
            schemes = result['schemes']
        else:
            schemes = []
            flash('Error loading schemes', 'error')
        
        # Get departments for dropdown
        dept_result = Department.get_all_departments(mongo, active_only=True)
        departments = dept_result['departments'] if dept_result['success'] else []
        
        # Get dashboard statistics for the template
        try:
            stats = PanchayatRecord.get_statistics(mongo)
            approved_stats = calculate_approved_stats(mongo)
            pending_stats = calculate_pending_stats(mongo)
            inreview_stats = calculate_inreview_stats(mongo)
            disapproved_stats = calculate_disapproved_stats(mongo)
            rejected_stats = calculate_rejected_stats(mongo)
        except Exception as e:
            print(f"Error loading dashboard stats: {e}")
            # Provide default values
            stats = {'total_records': 0, 'total_amount': 0, 'category_stats': [], 'panchayat_stats': []}
            approved_stats = {'count': 0, 'recent_count': 0}
            pending_stats = {'count': 0, 'recent_count': 0}
            inreview_stats = {'count': 0, 'recent_count': 0}
            disapproved_stats = {'count': 0, 'recent_count': 0}
            rejected_stats = {'count': 0, 'recent_count': 0}
        
        return render_template('admin/manage_schemes.html', 
                             schemes=schemes, 
                             departments=departments,
                             approved_status=approved_stats,
                             approved_stats=approved_stats,
                             pending_stats=pending_stats,
                             inreview_stats=inreview_stats,
                             disapproved_stats=disapproved_stats,
                             rejected_stats=rejected_stats,
                             records_count=stats['total_records'],
                             total_amount=stats['total_amount'])
        
    except Exception as e:
        flash('Error loading schemes', 'error')
        return render_template('admin/manage_schemes.html', 
                             schemes=[], 
                             departments=[],
                             approved_status={'count': 0, 'recent_count': 0},
                             approved_stats={'count': 0, 'recent_count': 0},
                             pending_stats={'count': 0, 'recent_count': 0},
                             inreview_stats={'count': 0, 'recent_count': 0},
                             disapproved_stats={'count': 0, 'recent_count': 0},
                             rejected_stats={'count': 0, 'recent_count': 0},
                             records_count=0,
                             total_amount=0)

@admin_bp.route('/create-scheme', methods=['POST'])
@login_required
@superadmin_required
def create_scheme():
    """Create a new scheme"""
    try:
        scheme_data = {
            'name': request.form.get('name', '').strip(),
            'department_id': request.form.get('department_id'),
            'description': request.form.get('description', '').strip(),
            'attributes': json.loads(request.form.get('attributes', '[]'))
        }
        
        if not scheme_data['name'] or not scheme_data['department_id']:
            flash('Scheme name and department are required', 'error')
            return redirect(url_for('admin.manage_schemes'))
        
        result = Scheme.create_scheme(mongo, scheme_data,username=session.get('username'))
        
        if result['success']:
            flash('Scheme created successfully!', 'success')
            # Redirect with the department ID to open its accordion
            return redirect(url_for('admin.manage_departments', open_department=scheme_data['department_id']))
        else:
            flash(result['message'], 'error')
            
    except Exception as e:
        flash('An error occurred while creating the scheme.', 'error')
    
    return redirect(url_for('admin.manage_departments'))

@admin_bp.route('/edit-scheme/<scheme_id>', methods=['GET', 'POST'])
@login_required
@superadmin_required
def edit_scheme(scheme_id):
    """Edit scheme"""
    if request.method == 'POST':
        try:
            update_data = {
                'name': request.form.get('name', '').strip(),
                'description': request.form.get('description', '').strip(),
                'attributes': json.loads(request.form.get('attributes', '[]'))
            }
            
            if not update_data['name']:
                flash('Scheme name is required', 'error')
                return redirect(url_for('admin.manage_schemes'))
            
            result = Scheme.update_scheme(mongo, scheme_id, update_data,username=session.get('username'))
            
            if result['success']:
                flash('Scheme updated successfully!', 'success')
            else:
                flash(result['message'], 'error')
                
        except Exception as e:
            flash('An error occurred while updating the scheme.', 'error')
        
        return redirect(url_for('admin.manage_departments'))
    
    # GET request - show edit form
    try:
        scheme = Scheme.get_scheme_by_id(mongo, scheme_id)
        if not scheme:
            flash('Scheme not found', 'error')
            return redirect(url_for('admin.manage_departments'))
        
        # Get departments for dropdown
        dept_result = Department.get_all_departments(mongo, active_only=True)
        departments = dept_result['departments'] if dept_result['success'] else []
        
        return render_template('admin/edit_scheme.html', scheme=scheme, departments=departments)
        
    except Exception as e:
        flash('Error loading scheme', 'error')
        return redirect(url_for('admin.manage_departments'))

@admin_bp.route('/delete-scheme/<scheme_id>')
@login_required
@superadmin_required
def delete_scheme(scheme_id):
    """Delete scheme"""
    try:
        result = Scheme.delete_scheme(mongo, scheme_id,username=session.get('username'))
        
        if result['success']:
            flash('Scheme deleted successfully!', 'success')
        else:
            flash(result['message'], 'error')
            
    except Exception as e:
        flash('Error deleting scheme', 'error')
    
    return redirect(url_for('admin.manage_departments'))

# API Routes for AJAX calls

@admin_bp.route('/api/schemes-by-department/<department_id>')
@login_required
@admin_required
def get_schemes_by_department(department_id):
    """Get schemes for a specific department"""
    try:
        result = Scheme.get_schemes_by_department(mongo, department_id)
        if result['success']:
            return jsonify({'success': True, 'schemes': result['schemes']})
        else:
            return jsonify({'success': False, 'message': result['message']})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@admin_bp.route('/api/scheme-attributes/<scheme_id>')
@login_required
@admin_required
def get_scheme_attributes(scheme_id):
    """Get attributes for a specific scheme"""
    try:
        result = Scheme.get_scheme_attributes(mongo, scheme_id)
        if result['success']:
            return jsonify({'success': True, 'attributes': result['attributes']})
        else:
            return jsonify({'success': False, 'message': result['message']})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@admin_bp.route('/api/data-types')
@login_required
@admin_required
def get_data_types():
    """Get available data types for scheme attributes"""
    try:
        data_types = Scheme.get_data_types()
        return jsonify({'success': True, 'data_types': data_types})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@admin_bp.route('/api/advanced-filter', methods=['POST'])
@login_required
@admin_required
def advanced_filter():
    """Apply advanced filters to records"""
    try:
        data = request.get_json()
        page = data.get('page', 1)
        per_page = data.get('per_page', 10)
        search = data.get('search', '')
        department_ids = data.get('department_ids', [])
        scheme_ids = data.get('scheme_ids', [])
        filters = data.get('filters', [])
        
        # Get user's access information
        user_id = session.get('user_id')
        
        # Use PanchayatRecord model to get records with advanced filtering
        result = PanchayatRecord.get_all_records(
            mongo,
            page=page,
            per_page=per_page,
            search=search, 
            department_ids=department_ids if department_ids else None, 
            scheme_ids=scheme_ids if scheme_ids else None,
            filters=filters if filters else None,
            user_id=user_id
        )
        
        return jsonify(result)
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})


@admin_bp.route('/api/departments-for-user-creation')
@login_required
@superadmin_required
def get_departments_for_user_creation():
    """Get all departments for user creation form"""
    try:
        result = Department.get_all_departments(mongo, active_only=True)
        if result['success']:
            departments = []
            for dept in result['departments']:
                departments.append({
                    'id': str(dept['_id']),
                    'name': dept['name'],
                    'description': dept.get('description', '')
                })
            return jsonify({'success': True, 'departments': departments})
        else:
            return jsonify({'success': False, 'message': result['message']})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@admin_bp.route('/api/schemes-for-user-creation')
@login_required
@admin_required
def get_schemes_for_user_creation():
    """API endpoint to get schemes available for the logged-in user to create records"""
    try:
        user_id = session.get('user_id')
        user_access = User.get_user_access_info(mongo, user_id)
        
        if not user_access:
            return jsonify({
                'success': False,
                'message': 'User access information not found'
            })
        
        schemes = []
        
        # Check if user has all access or is superadmin
        if user_access.get('has_all_access') or user_access.get('is_superadmin'):
            # Get all active schemes
            schemes_result = Scheme.get_all_schemes(mongo, active_only=True)
            if schemes_result['success']:
                schemes = schemes_result['schemes']
        else:
            # Get only schemes the user has access to
            scheme_access = user_access.get('scheme_access', [])
            
            if scheme_access and 'all' not in scheme_access:
                try:
                    # Convert scheme IDs to ObjectId
                    scheme_object_ids = [ObjectId(s) for s in scheme_access if s != 'all']
                    
                    # Fetch schemes from database
                    schemes = list(mongo.db.schemes.find({
                        '_id': {'$in': scheme_object_ids},
                        'is_active': True
                    }).sort('name', 1))
                    
                except Exception as e:
                    print(f"ERROR SCHEMES_FOR_CREATION - Converting scheme IDs: {e}")
                    import traceback
                    traceback.print_exc()
                    schemes = []
            elif 'all' in scheme_access:
                # User has 'all' access - get all schemes
                schemes_result = Scheme.get_all_schemes(mongo, active_only=True)
                if schemes_result['success']:
                    schemes = schemes_result['schemes']
        
        # Get department information for each scheme
        schemes_with_dept = []
        for scheme in schemes:
            dept = mongo.db.departments.find_one({'_id': scheme.get('department_id')})
            
            schemes_with_dept.append({
                'id': str(scheme['_id']),
                'name': scheme['name'],
                'description': scheme.get('description', ''),
                'department_id': str(scheme.get('department_id')),
                'department_name': dept['name'] if dept else 'Unknown Department',
                'attributes': scheme.get('attributes', [])
            })
        
        return jsonify({
            'success': True,
            'schemes': schemes_with_dept
        })
        
    except Exception as e:
        print(f"ERROR in get_schemes_for_user_creation: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'Error fetching schemes: {str(e)}'
        })

@admin_bp.route('/api/superadmin-filter-data')
@login_required
@admin_required  # CHANGED from @superadmin_required to allow all admins
def get_superadmin_filter_data():
    """API endpoint to get all departments and schemes for filter dropdowns"""
    try:
        user_id = session.get('user_id')
        user_access = User.get_user_access_info(mongo, user_id)
        
        if not user_access:
            return jsonify({'success': False, 'message': 'User access info not found'})
        
        # Get departments based on user access
        if user_access.get('has_all_access') or user_access.get('is_superadmin'):
            departments_result = Department.get_all_departments(mongo, active_only=True)
            departments = departments_result['departments'] if departments_result['success'] else []
        else:
            dept_access = user_access.get('department_access', [])
            if dept_access and 'all' not in dept_access:
                departments = list(mongo.db.departments.find({
                    '_id': {'$in': [ObjectId(d) for d in dept_access if d != 'all']},
                    'is_active': True
                }).sort('name', 1))
            else:
                departments = []
        
        # Get schemes based on user access
        if user_access.get('has_all_access') or user_access.get('is_superadmin'):
            schemes_result = Scheme.get_all_schemes(mongo, active_only=True)
            schemes = schemes_result['schemes'] if schemes_result['success'] else []
        else:
            scheme_access = user_access.get('scheme_access', [])
            if scheme_access and 'all' not in scheme_access:
                schemes = list(mongo.db.schemes.find({
                    '_id': {'$in': [ObjectId(s) for s in scheme_access if s != 'all']},
                    'is_active': True
                }).sort('name', 1))
            else:
                schemes = []
        
        # Add department names to schemes
        for scheme in schemes:
            dept = mongo.db.departments.find_one({'_id': scheme.get('department_id')})
            scheme['department_name'] = dept['name'] if dept else 'Unknown'
        
        # Convert ObjectId to string for JSON serialization
        departments_list = []
        for dept in departments:
            departments_list.append({
                '_id': {'$oid': str(dept['_id'])},
                'name': dept['name']
            })
        
        schemes_list = []
        for scheme in schemes:
            schemes_list.append({
                '_id': {'$oid': str(scheme['_id'])},
                'name': scheme['name'],
                'department_id': {'$oid': str(scheme['department_id'])},
                'department_name': scheme.get('department_name', 'Unknown')
            })
        
        return jsonify({
            'success': True,
            'departments': departments_list,
            'schemes': schemes_list
        })
        
    except Exception as e:
        print(f"Error in get_superadmin_filter_data: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'Error fetching filter data: {str(e)}'
        })

@admin_bp.route('/api/record-details/<record_id>')
@login_required
@admin_required
def get_record_details(record_id):
    """API endpoint to get record details for editing"""
    try:
        from bson import ObjectId
        
        # Get record by ID
        record = mongo.db.panchayat_records.find_one({'_id': ObjectId(record_id)})
        
        if not record:
            return jsonify({'success': False, 'message': 'Record not found'}), 404
        
        # Convert ObjectId to string for JSON serialization
        record['_id'] = {'$oid': str(record['_id'])}
        if 'department_id' in record:
            record['department_id'] = {'$oid': str(record['department_id'])}
        if 'scheme_id' in record:
            record['scheme_id'] = {'$oid': str(record['scheme_id'])}
        
        return jsonify({
            'success': True,
            'record': record
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@admin_bp.route('/bulk-import')
@login_required
@admin_required
def bulk_import():
    """Bulk import page"""
    return render_template('admin/bulk_import.html')

@admin_bp.route('/import-history')
@login_required
@admin_required
def import_history():
    """Import history page"""
    return render_template('admin/import_history.html')

@admin_bp.route('/api/start-bulk-import', methods=['POST'])
@login_required
@admin_required
def start_bulk_import():
    """Start bulk import process"""
    try:
        from models.import_history import ImportHistory
        import threading
        import os
        
        # Get form data
        file = request.files.get('file')
        scheme_id = request.form.get('scheme_id')
        department_id = request.form.get('department_id')
        skip_duplicates = request.form.get('skip_duplicates') == 'on'
        user_id = session.get('user_id')
        
        if not file or not scheme_id or not department_id:
            return jsonify({'success': False, 'message': 'Missing required fields'}), 400
        
        # Validate file
        if not file.filename:
            return jsonify({'success': False, 'message': 'No file selected'}), 400
        
        # Check file extension
        allowed_extensions = {'.xlsx', '.xls', '.csv'}
        file_ext = os.path.splitext(file.filename)[1].lower()
        if file_ext not in allowed_extensions:
            return jsonify({'success': False, 'message': 'Invalid file format. Only Excel (.xlsx, .xls) or CSV (.csv) files are allowed.'}), 400
        
        # Check file size (10MB limit)
        file.seek(0, 2)  # Seek to end
        file_size = file.tell()
        file.seek(0)  # Reset to beginning
        if file_size > 10 * 1024 * 1024:  # 10MB
            return jsonify({'success': False, 'message': 'File size exceeds 10MB limit'}), 400
        
        # Get user info
        user = mongo.db.users.find_one({'_id': ObjectId(user_id)})
        if not user:
            return jsonify({'success': False, 'message': 'User not found'}), 400
        
        # Create import record
        import_result = ImportHistory.create_import_record(
            mongo=mongo,
            file_name=file.filename,
            scheme_id=scheme_id,
            department_id=department_id,
            imported_by=user.get('username', 'Unknown')
        )
        
        if not import_result['success']:
            return jsonify(import_result), 400
        
        import_id = import_result['import_id']
        
        # Save file temporarily
        import tempfile
        temp_dir = tempfile.mkdtemp()
        temp_file_path = os.path.join(temp_dir, file.filename)
        file.save(temp_file_path)
        
        # Start background import process
        import_thread = threading.Thread(
            target=process_bulk_import,
            args=(import_id, temp_file_path, scheme_id, department_id, skip_duplicates, user.get('username'))
        )
        import_thread.daemon = True
        import_thread.start()
        
        return jsonify({
            'success': True,
            'message': 'Import started successfully',
            'import_id': import_id
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@admin_bp.route('/api/import-history')
@login_required
@admin_required
def get_import_history():
    """Get import history"""
    try:
        from models.import_history import ImportHistory
        
        user_id = session.get('user_id')
        user = mongo.db.users.find_one({'_id': ObjectId(user_id)})
        username = user.get('username', 'Unknown') if user else 'Unknown'
        is_superadmin = user.get('is_superadmin', False) if user else False
        
        # Superadmins can see all imports, regular users see only their own
        if is_superadmin:
            result = ImportHistory.get_import_history(mongo, username=None, limit=100)
        else:
            result = ImportHistory.get_import_history(mongo, username=username, limit=100)
        
        return jsonify(result)
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@admin_bp.route('/api/import-details/<import_id>')
@login_required
@admin_required
def get_import_details(import_id):
    """Get import details by ID"""
    try:
        from models.import_history import ImportHistory
        
        result = ImportHistory.get_import_by_id(mongo, import_id)
        return jsonify(result)
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@admin_bp.route('/api/stop-import/<import_id>', methods=['POST'])
@login_required
@admin_required
def stop_import(import_id):
    """Stop an import operation"""
    try:
        from models.import_history import ImportHistory
        
        user_id = session.get('user_id')
        user = mongo.db.users.find_one({'_id': ObjectId(user_id)})
        stopped_by = user.get('username', 'Unknown') if user else 'Unknown'
        
        result = ImportHistory.stop_import(mongo, import_id, stopped_by)
        return jsonify(result)
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@admin_bp.route('/api/download-template/<scheme_id>')
@login_required
@admin_required
def download_template(scheme_id):
    """Download import template for a scheme"""
    try:
        from models.scheme import Scheme
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        from io import BytesIO
        
        # Get scheme details
        scheme = mongo.db.schemes.find_one({'_id': ObjectId(scheme_id)})
        if not scheme:
            return jsonify({'success': False, 'message': 'Scheme not found'}), 404
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Import Template"
        
        # Add headers - only scheme-specific fields
        headers = []
        
        # Only include scheme-specific attributes
        if scheme.get('attributes'):
            for attr in scheme['attributes']:
                if attr.get('name'):
                    headers.append(attr['name'])
        
        # If no attributes defined, use basic required fields
        if not headers:
            headers = ['beneficiary_name', 'father_name', 'mother_name', 'village_name', 'registration_number']
        
        # Style for headers
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header.replace('_', ' ').title())
            cell.font = header_font
            cell.fill = header_fill
        
        # Add sample data row
        sample_data = []
        for header in headers:
            if 'name' in header.lower():
                sample_data.append('Sample Name')
            elif 'father' in header.lower():
                sample_data.append('Sample Father')
            elif 'mother' in header.lower():
                sample_data.append('Sample Mother')
            elif 'village' in header.lower():
                sample_data.append('Sample Village')
            elif 'registration' in header.lower():
                sample_data.append('Sample Reg No')
            elif 'income' in header.lower():
                sample_data.append('50000')
            elif 'size' in header.lower():
                sample_data.append('4')
            elif 'type' in header.lower():
                sample_data.append('Sample Type')
            else:
                sample_data.append('Sample Value')
        
        for col, data in enumerate(sample_data, 1):
            ws.cell(row=2, column=col, value=data)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Return file
        from flask import send_file
        return send_file(
            output,
            as_attachment=True,
            download_name=f"{scheme['name']}_template.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@admin_bp.route('/api/download-template-csv/<scheme_id>')
@login_required
@admin_required
def download_template_csv(scheme_id):
    """Download CSV import template for a scheme"""
    try:
        import csv
        from io import StringIO
        
        # Get scheme details
        scheme = mongo.db.schemes.find_one({'_id': ObjectId(scheme_id)})
        if not scheme:
            return jsonify({'success': False, 'message': 'Scheme not found'}), 404
        
        # Create CSV content
        output = StringIO()
        writer = csv.writer(output)
        
        # Add headers - only scheme-specific fields
        headers = []
        
        # Only include scheme-specific attributes
        if scheme.get('attributes'):
            for attr in scheme['attributes']:
                if attr.get('name'):
                    headers.append(attr['name'])
        
        # If no attributes defined, use basic required fields
        if not headers:
            headers = ['beneficiary_name', 'father_name', 'mother_name', 'village_name', 'registration_number']
        
        # Write headers
        writer.writerow([header.replace('_', ' ').title() for header in headers])
        
        # Add sample data row
        sample_data = []
        for header in headers:
            if 'name' in header.lower():
                sample_data.append('Sample Name')
            elif 'father' in header.lower():
                sample_data.append('Sample Father')
            elif 'mother' in header.lower():
                sample_data.append('Sample Mother')
            elif 'village' in header.lower():
                sample_data.append('Sample Village')
            elif 'registration' in header.lower():
                sample_data.append('Sample Reg No')
            elif 'income' in header.lower():
                sample_data.append('50000')
            elif 'size' in header.lower():
                sample_data.append('4')
            elif 'type' in header.lower():
                sample_data.append('Sample Type')
            else:
                sample_data.append('Sample Value')
        
        writer.writerow(sample_data)
        
        # Get CSV content
        csv_content = output.getvalue()
        output.close()
        
        # Return file
        from flask import Response
        return Response(
            csv_content,
            mimetype='text/csv',
            headers={
                'Content-Disposition': f'attachment; filename="{scheme["name"]}_template.csv"'
            }
        )
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@admin_bp.route('/api/validate-import-file', methods=['POST'])
@login_required
@admin_required
def validate_import_file():
    """Validate import file against scheme structure"""
    try:
        import pandas as pd
        import os
        
        # Get form data
        file = request.files.get('file')
        scheme_id = request.form.get('scheme_id')
        
        if not file or not scheme_id:
            return jsonify({'success': False, 'message': 'Missing required fields'}), 400
        
        # Get scheme details
        scheme = mongo.db.schemes.find_one({'_id': ObjectId(scheme_id)})
        if not scheme:
            return jsonify({'success': False, 'message': 'Scheme not found'}), 404
        
        # Read file
        try:
            file_ext = os.path.splitext(file.filename)[1].lower()
            if file_ext == '.csv':
                df = pd.read_csv(file)
            else:
                df = pd.read_excel(file)
        except Exception as e:
            return jsonify({'success': False, 'message': f'Error reading file: {str(e)}'}), 400
        
        # Validate file structure
        validation_result = validate_file_structure(df, scheme)
        
        return jsonify({
            'success': True,
            'validation': validation_result
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

def validate_file_structure(df, scheme):
    """Validate file structure against scheme requirements"""
    errors = []
    
    # Get required columns ONLY from scheme attributes
    required_columns = []
    
    # Only validate against scheme-specific attributes
    if scheme.get('attributes'):
        for attr in scheme['attributes']:
            if attr.get('name'):
                required_columns.append(attr['name'])
    
    # If no attributes defined in scheme, use basic required fields
    if not required_columns:
        required_columns = ['beneficiary_name', 'father_name', 'mother_name', 'village_name', 'registration_number']
    
    # Check if file has data
    if len(df) == 0:
        errors.append('File is empty - no data rows found')
        return {'valid': False, 'errors': errors}
    
    # Get file columns (normalize to lowercase)
    file_columns = [col.lower().strip().replace(' ', '_') for col in df.columns]
    
    # Check for missing required columns
    missing_columns = []
    for required_col in required_columns:
        if required_col not in file_columns:
            missing_columns.append(required_col.replace('_', ' ').title())
    
    if missing_columns:
        errors.append(f'Missing required columns: {", ".join(missing_columns)}')
    
    # Check for extra columns (warn but don't fail)
    extra_columns = []
    for file_col in file_columns:
        if file_col not in required_columns:
            extra_columns.append(file_col.replace('_', ' ').title())
    
    # Check for empty required fields
    empty_required_fields = []
    for required_col in required_columns:
        if required_col in file_columns:
            # Check if any values in this column are empty/null
            if df[required_col].isna().all() or (df[required_col] == '').all():
                empty_required_fields.append(required_col.replace('_', ' ').title())
    
    if empty_required_fields:
        errors.append(f'Required columns are completely empty: {", ".join(empty_required_fields)}')
    
    # Check data types for numeric fields
    type_errors = []
    if scheme.get('attributes'):
        for attr in scheme['attributes']:
            attr_name = attr.get('name')
            attr_type = attr.get('type')
            
            if attr_name in file_columns and attr_type in ['int', 'float']:
                # Try to convert to numeric, check for errors
                try:
                    pd.to_numeric(df[attr_name], errors='coerce')
                    # Check if all values became NaN (meaning conversion failed)
                    if pd.to_numeric(df[attr_name], errors='coerce').isna().all():
                        type_errors.append(f'Column "{attr_name.replace("_", " ").title()}" should contain {attr_type} values')
                except:
                    type_errors.append(f'Column "{attr_name.replace("_", " ").title()}" should contain {attr_type} values')
    
    if type_errors:
        errors.extend(type_errors)
    
    # Return validation result
    is_valid = len(errors) == 0
    
    result = {
        'valid': is_valid,
        'errors': errors,
        'record_count': len(df),
        'required_columns': [col.replace('_', ' ').title() for col in required_columns],
        'file_columns': [col.replace('_', ' ').title() for col in file_columns],
        'extra_columns': [col.replace('_', ' ').title() for col in extra_columns] if extra_columns else []
    }
    
    return result

def process_bulk_import(import_id, file_path, scheme_id, department_id, skip_duplicates, username):
    """Background process for bulk import"""
    try:
        from models.import_history import ImportHistory, ImportStatus
        from models.admin import PanchayatRecord
        import pandas as pd
        import os
        import shutil
        
        # Update status to in progress
        ImportHistory.update_import_status(mongo, import_id, ImportStatus.IN_PROGRESS)
        
        # Read file (Excel or CSV)
        try:
            file_ext = os.path.splitext(file_path)[1].lower()
            if file_ext == '.csv':
                df = pd.read_csv(file_path)
            else:
                df = pd.read_excel(file_path)
        except Exception as e:
            ImportHistory.update_import_status(
                mongo, import_id, ImportStatus.FAILED,
                error_message=f"Error reading file: {str(e)}"
            )
            return
        
        # Get total records
        total_records = len(df)
        total_batches = (total_records + 99) // 100  # Batch size of 100
        
        # Update total records and batches
        mongo.db.import_history.update_one(
            {'_id': ObjectId(import_id)},
            {'$set': {
                'total_records': total_records,
                'total_batches': total_batches
            }}
        )
        
        # Process in batches
        imported_count = 0
        failed_count = 0
        duplicate_count = 0
        current_batch = 0
        
        for batch_start in range(0, total_records, 100):
            # Check if import was stopped
            import_record = mongo.db.import_history.find_one({'_id': ObjectId(import_id)})
            if import_record and import_record.get('status') == ImportStatus.STOPPED.value:
                break
            
            current_batch += 1
            batch_end = min(batch_start + 100, total_records)
            batch_df = df.iloc[batch_start:batch_end]
            
            # Process batch
            batch_imported = 0
            batch_failed = 0
            batch_duplicate = 0
            
            for index, row in batch_df.iterrows():
                try:
                    # Check if import was stopped
                    import_record = mongo.db.import_history.find_one({'_id': ObjectId(import_id)})
                    if import_record and import_record.get('status') == ImportStatus.STOPPED.value:
                        break
                    
                    # Prepare record data
                    record_data = {
                        'department_id': ObjectId(department_id),
                        'scheme_id': ObjectId(scheme_id),
                        'created_by': username,
                        'custom_data': {}
                    }
                    
                    # Map Excel columns to record fields
                    for column in df.columns:
                        value = row[column]
                        if pd.notna(value):
                            # Convert to appropriate type
                            if isinstance(value, (int, float)):
                                record_data['custom_data'][column.lower().replace(' ', '_')] = value
                            else:
                                record_data['custom_data'][column.lower().replace(' ', '_')] = str(value)
                    
                    # Check for duplicates if skip_duplicates is enabled
                    if skip_duplicates:
                        # Simple duplicate check based on beneficiary name and registration number
                        beneficiary_name = record_data['custom_data'].get('beneficiary_name', '')
                        registration_number = record_data['custom_data'].get('registration_number', '')
                        
                        if beneficiary_name and registration_number:
                            existing = mongo.db.panchayat_records.find_one({
                                'custom_data.beneficiary_name': beneficiary_name,
                                'custom_data.registration_number': registration_number,
                                'scheme_id': ObjectId(scheme_id),
                                'is_active': True
                            })
                            
                            if existing:
                                batch_duplicate += 1
                                continue
                    
                    # Create record
                    result = PanchayatRecord.create_record(mongo, record_data, username)
                    if result['success']:
                        batch_imported += 1
                    else:
                        batch_failed += 1
                        
                except Exception as e:
                    batch_failed += 1
                    print(f"Error processing row {index}: {str(e)}")
            
            # Update counts
            imported_count += batch_imported
            failed_count += batch_failed
            duplicate_count += batch_duplicate
            
            # Update progress
            ImportHistory.update_import_progress(
                mongo, import_id, imported_count, failed_count, duplicate_count, current_batch
            )
        
        # Check final status
        import_record = mongo.db.import_history.find_one({'_id': ObjectId(import_id)})
        if import_record and import_record.get('status') == ImportStatus.STOPPED.value:
            # Import was stopped, status already updated
            pass
        else:
            # Import completed
            ImportHistory.update_import_status(mongo, import_id, ImportStatus.COMPLETED)
        
    except Exception as e:
        # Update status to failed
        ImportHistory.update_import_status(
            mongo, import_id, ImportStatus.FAILED,
            error_message=f"Import failed: {str(e)}"
        )
        print(f"Bulk import failed: {str(e)}")
    
    finally:
        # Clean up temporary file
        try:
            if os.path.exists(file_path):
                os.remove(file_path)
            # Remove temp directory if empty
            temp_dir = os.path.dirname(file_path)
            if os.path.exists(temp_dir) and not os.listdir(temp_dir):
                os.rmdir(temp_dir)
        except Exception as e:
            print(f"Error cleaning up temp files: {str(e)}")

@admin_bp.route('/api/filtered-records')
@login_required
@admin_required
def get_filtered_records_api():
    """API endpoint to get filtered records for all admins with proper pagination"""
    try:
        # Get filter parameters from query string
        department_id = request.args.get('department_id')
        scheme_id = request.args.get('scheme_id')
        taluka = request.args.get('taluka')
        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 10))
        search = request.args.get('search', '')
        
        # Get user ID for access filtering
        user_id = session.get('user_id')
        
        # Convert to lists for the get_all_records method
        department_ids = [department_id] if department_id else None
        scheme_ids = [scheme_id] if scheme_id else None
        
        # Get filtered records with user access control and proper pagination
        result = PanchayatRecord.get_all_records(
            mongo, 
            page=page, 
            per_page=per_page,
            search=search,
            department_ids=department_ids,
            scheme_ids=scheme_ids,
            taluka_filter=taluka,
            user_id=user_id
        )
        
        if result['success']:
            # Convert ObjectId to string for JSON serialization
            records = result['records']
            for record in records:
                if '_id' in record:
                    record['_id'] = {'$oid': str(record['_id'])} if not isinstance(record['_id'], dict) else record['_id']
                if 'department_id' in record:
                    record['department_id'] = {'$oid': str(record['department_id'])} if not isinstance(record['department_id'], dict) else record['department_id']
                if 'scheme_id' in record:
                    record['scheme_id'] = {'$oid': str(record['scheme_id'])} if not isinstance(record['scheme_id'], dict) else record['scheme_id']
            
            return jsonify({
                'success': True,
                'records': records,
                'total_records': result['total_records'],
                'total_pages': result['total_pages'],
                'page': result['page'],
                'per_page': result['per_page']
            })
        else:
            return jsonify({
                'success': False,
                'message': result.get('message', 'Unknown error')
            })
        
    except Exception as e:
        print(f"ERROR in get_filtered_records_api: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'Error fetching filtered records: {str(e)}'
        })

@admin_bp.route('/api/dashboard-data')
@login_required
@admin_required
def get_dashboard_data_api():
    """API endpoint to get filtered data for smart dashboard"""
    try:
        # Get filter parameters from query string
        department_id = request.args.get('department_id')
        scheme_id = request.args.get('scheme_id')
        taluka = request.args.get('taluka')
        
        # Get user ID for access filtering
        user_id = session.get('user_id')
        
        # Convert to lists for the get_all_records method
        department_ids = [department_id] if department_id else None
        scheme_ids = [scheme_id] if scheme_id else None
        
        # Get filtered records with user access control
        result = PanchayatRecord.get_all_records(
            mongo, 
            page=1, 
            per_page=100000,  # Get all records for dashboard
            department_ids=department_ids,
            scheme_ids=scheme_ids,
            taluka_filter=taluka,
            user_id=user_id  # IMPORTANT: Pass user_id for access filtering
        )
        
        if result['success']:
            # Convert ObjectId to string for JSON serialization
            records = result['records']
            for record in records:
                if '_id' in record:
                    record['_id'] = str(record['_id'])
                if 'department_id' in record:
                    record['department_id'] = str(record['department_id'])
                if 'scheme_id' in record:
                    record['scheme_id'] = str(record['scheme_id'])
            
            return jsonify({
                'success': True,
                'records': records,
                'total_records': result['total_records']
            })
        else:
            return jsonify({
                'success': False,
                'message': result.get('message', 'Unknown error')
            })
        
    except Exception as e:
        print(f"Error in get_dashboard_data_api: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'Error fetching dashboard data: {str(e)}'
        })

@admin_bp.route('/api/download-records')
@login_required
@superadmin_required
def download_records_api():
    """API endpoint to download filtered records as Excel"""
    try:
        # Get filter parameters from query string
        department_id = request.args.get('department_id')
        scheme_id = request.args.get('scheme_id')
        taluka = request.args.get('taluka')
        
        # Convert to lists for the get_all_records method
        department_ids = [department_id] if department_id else None
        scheme_ids = [scheme_id] if scheme_id else None
        
        result = PanchayatRecord.get_all_records(
            mongo, 
            page=1, 
            per_page=10000,
            department_ids=department_ids,
            scheme_ids=scheme_ids,
            taluka_filter=taluka
        )
        
        if not result['success']:
            return jsonify({'success': False, 'message': result['message']})
        
        records = result['records']
        
        # Create Excel file
        import pandas as pd
        from io import BytesIO
        
        # Prepare data for Excel
        excel_data = []
        for record in records:
            row_data = {
                'Department': record.get('department_name', ''),
                'Scheme': record.get('scheme_name', ''),
                'Created By': record.get('created_by', ''),
                'Created At': record.get('created_at', ''),
                'Status': 'Active' if record.get('is_active', True) else 'Inactive'
            }
            
            # Add custom_data fields
            if record.get('custom_data'):
                for key, value in record['custom_data'].items():
                    row_data[key.replace('_', ' ').title()] = value
            
            excel_data.append(row_data)
        
        # Create DataFrame
        df = pd.DataFrame(excel_data)
        
        # Create Excel file in memory
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Records', index=False)
        
        output.seek(0)
        
        # Create response
        from flask import Response
        response = Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': 'attachment; filename=filtered_records.xlsx'}
        )
        
        return response
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error creating download: {str(e)}'
        })

@admin_bp.route('/api/scheme-form-fields/<scheme_id>')
@login_required
@admin_required
def get_scheme_form_fields(scheme_id):
    """Get form fields for a specific scheme"""
    try:
        # Get scheme details
        scheme_result = Scheme.get_scheme_by_id(mongo, scheme_id)
        
        if not scheme_result['success']:
            return jsonify({'success': False, 'message': 'Scheme not found'})
        
        scheme = scheme_result['scheme']
        attributes = scheme.get('attributes', [])
        
        # Convert attributes to form fields
        form_fields = []
        for attr in attributes:
            field = {
                'name': attr.get('name', ''),
                'label': attr.get('label', ''),
                'type': attr.get('type', 'string'),
                'required': attr.get('required', False),
                'options': attr.get('options', []) if attr.get('type') == 'enum' else []
            }
            form_fields.append(field)
        
        return jsonify({
            'success': True,
            'scheme_name': scheme['name'],
            'department_name': scheme.get('department_name', 'Unknown'),
            'form_fields': form_fields
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})
